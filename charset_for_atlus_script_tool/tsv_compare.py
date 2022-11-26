import os
import openpyxl

EMPTYPLACEHOLDER = "*"

def readTsv(filePath):
  with open(filePath, 'r',encoding='utf8') as file:
    data = []
    lines = file.read().splitlines()
    for line in lines:
      data.append(line.split('\t'))
    return data

def checkTsvShape(data):
  keys = list(data.keys())
  if len(keys) == 0 :
    raise Exception("empty table")
  # width, height
  shape = (len(tsvs[keys[0]][0]), len(tsvs[keys[0]]))
  for fileName in keys[1:]:
    if not shape == (len(tsvs[fileName][0]), len(tsvs[fileName])):
      raise Exception("table shape not correspond")
  return shape

def isAllTheSame(data):
  if data.count(EMPTYPLACEHOLDER) == len(data):
    return False
  return data.count(data[0]) == len(data)

def corespdTabXlsx(table, header):
  keys = list(table.keys())
  sum = len(keys)

  #gather and calulate empty rate
  values = list(table.values())
  emptyRates = {}
  for index in range(len(header)):
    column = [x[index] for x in values]
    emptyRates[header[index]] = column.count(EMPTYPLACEHOLDER)/sum
  workBook = openpyxl.Workbook()
  workSheet=workBook.create_sheet('文件名(空字率)',0)
  workSheet.cell(row = 1, column =1, value = "行,列")
  for fileName in header:
    workSheet.cell(row=1, column = header.index(fileName) + 2, value="\t{} ({:.2f}%)\t".format(fileName, emptyRates[fileName]*100))

  for index in range(sum):
    workSheet.cell(row=index + 2, column = 1, value = "{}, {}".format(keys[index][0],keys[index][1]))
    for j in range(len(values[index])):
      workSheet.cell(row= index + 2, column = j + 2, value=values[index][j])
  workBook.save("tsv_result_compare.xlsx")
  
def corespdTabTsv(table, header):
  emptyPlaceHolder = "*"
  keys = list(table.keys())
  sum = len(keys)

  #gather and calulate empty rate
  values = list(table.values())
  emptyRates = {}
  for index in range(len(header)):
    column = [x[index] for x in values]
    emptyRates[header[index]] = column.count(emptyPlaceHolder)/len(keys)
  with open('tsv_result_compare.tsv','w', encoding="utf8") as file: 
    file.write("文件名(空字率)\n")
    # print header
    line = ""
    for fileName in header:
      line += "\t{} ({:.2f}%)\t".format(fileName, emptyRates[fileName]*100)
    line += "\n"
    file.write(line)
    # print not Corespd values
    for index in range(sum):
      line = "{}\t".format(keys[index])
      for char in values[index]:
     
        line += "{}\t".format(char)
      line+="\n"
      file.write(line)

if __name__ == '__main__':
  tsvFolder = "tsv"
  tsvFileNames = os.listdir(tsvFolder)
  tsvs = {}
  for fileNmae in tsvFileNames:
    tsv = readTsv(os.path.join(tsvFolder, fileNmae))
    tsvs[fileNmae] = tsv
  shape = checkTsvShape(tsvs)

  correspondingTable = {}
  for row in range(shape[1]):
    for column in range(shape[0]):
      correspond = []

      # correspondingTableHeader
      for tsv in tsvFileNames:
        correspond.append(tsvs[tsv][row][column])
      if(not isAllTheSame(correspond)):
        correspondingTable[(row,column)] = correspond
  # corespdTabTsv(correspondingTable, tsvFileNames)
  corespdTabXlsx(correspondingTable, tsvFileNames)
